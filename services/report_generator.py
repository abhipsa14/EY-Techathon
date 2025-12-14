"""
Report Generator Service - Creates PDF and CSV reports for validation results.
Generates detailed reports, summaries, and exports.
"""

import csv
import io
from typing import List, Dict, Any, Optional
from datetime import datetime
from pathlib import Path

from models import (
    Provider, ValidationResult, ValidationReport, Discrepancy,
    ValidationStatus, Priority
)
from config import REPORTS_DIR


class ReportGenerator:
    """Service for generating validation reports in various formats."""
    
    def __init__(self):
        self.reports_dir = REPORTS_DIR
        self.reports_dir.mkdir(exist_ok=True)
        
    def generate_csv_report(
        self, 
        providers: List[Provider],
        validation_results: Dict[str, ValidationResult],
        filename: Optional[str] = None
    ) -> str:
        """
        Generate a CSV report of validation results.
        
        Returns:
            Path to the generated CSV file
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"validation_report_{timestamp}.csv"
        
        filepath = self.reports_dir / filename
        
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = [
                'NPI', 'Provider Name', 'Practice Name', 'Specialty',
                'Phone', 'Address', 'City', 'State', 'Zip',
                'Status', 'Confidence Score', 'Discrepancies',
                'Auto Updated', 'Needs Review', 'Urgent Review',
                'Validated At'
            ]
            
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            
            for provider in providers:
                result = validation_results.get(provider.id)
                
                row = {
                    'NPI': provider.npi,
                    'Provider Name': provider.full_name(),
                    'Practice Name': provider.practice_name,
                    'Specialty': provider.specialty,
                    'Phone': provider.contact.phone,
                    'Address': provider.address.street1,
                    'City': provider.address.city,
                    'State': provider.address.state,
                    'Zip': provider.address.zip_code,
                    'Status': result.status.value if result else 'pending',
                    'Confidence Score': f"{result.overall_confidence:.1f}%" if result else 'N/A',
                    'Discrepancies': result.total_discrepancies if result else 0,
                    'Auto Updated': 'Yes' if result and result.auto_updated else 'No',
                    'Needs Review': 'Yes' if result and result.needs_review else 'No',
                    'Urgent Review': 'Yes' if result and result.urgent_review else 'No',
                    'Validated At': result.validated_at.strftime('%Y-%m-%d %H:%M') if result else 'N/A'
                }
                
                writer.writerow(row)
        
        return str(filepath)
    
    def generate_discrepancy_report(
        self,
        discrepancies: List[Discrepancy],
        providers: Dict[str, Provider],
        filename: Optional[str] = None
    ) -> str:
        """
        Generate a detailed discrepancy report in CSV format.
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"discrepancy_report_{timestamp}.csv"
        
        filepath = self.reports_dir / filename
        
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = [
                'Discrepancy ID', 'Provider NPI', 'Provider Name',
                'Type', 'Field', 'Current Value', 'Validated Value',
                'Source', 'Priority', 'Confidence', 'Detected At', 'Resolved'
            ]
            
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            
            for disc in discrepancies:
                provider = providers.get(disc.provider_id)
                
                row = {
                    'Discrepancy ID': disc.id[:8],
                    'Provider NPI': provider.npi if provider else 'Unknown',
                    'Provider Name': provider.full_name() if provider else 'Unknown',
                    'Type': disc.type.value.replace('_', ' ').title(),
                    'Field': disc.field_name,
                    'Current Value': disc.current_value,
                    'Validated Value': disc.validated_value,
                    'Source': disc.source.value,
                    'Priority': disc.priority.value.upper(),
                    'Confidence': f"{disc.confidence:.1f}%",
                    'Detected At': disc.detected_at.strftime('%Y-%m-%d %H:%M'),
                    'Resolved': 'Yes' if disc.resolved else 'No'
                }
                
                writer.writerow(row)
        
        return str(filepath)
    
    def generate_pdf_report(
        self,
        providers: List[Provider],
        validation_results: Dict[str, ValidationResult],
        report: ValidationReport,
        filename: Optional[str] = None
    ) -> str:
        """
        Generate a PDF summary report.
        
        Note: In production, this would use reportlab for actual PDF generation.
        For demo purposes, we generate an HTML file that can be printed to PDF.
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"validation_summary_{timestamp}.html"
        
        filepath = self.reports_dir / filename
        
        # Generate statistics
        status_counts = {
            'validated': sum(1 for p in providers if validation_results.get(p.id) and validation_results[p.id].auto_updated),
            'needs_review': sum(1 for p in providers if validation_results.get(p.id) and validation_results[p.id].needs_review),
            'urgent': sum(1 for p in providers if validation_results.get(p.id) and validation_results[p.id].urgent_review)
        }
        
        # Generate discrepancy breakdown
        disc_types = {}
        for p in providers:
            result = validation_results.get(p.id)
            if result:
                for disc in result.discrepancies:
                    dtype = disc.type.value
                    disc_types[dtype] = disc_types.get(dtype, 0) + 1
        
        # Build provider rows
        provider_rows = ""
        for p in providers[:50]:  # Limit to 50 for readability
            result = validation_results.get(p.id)
            if result:
                status_color = "#28a745" if result.auto_updated else ("#ffc107" if result.needs_review else "#dc3545")
                status_text = "✓" if result.auto_updated else ("⚠" if result.needs_review else "✗")
                
                provider_rows += f"""
                <tr>
                    <td>{p.npi}</td>
                    <td>{p.full_name()}</td>
                    <td>{p.specialty}</td>
                    <td style="text-align: center; color: {status_color}; font-weight: bold;">{status_text}</td>
                    <td style="text-align: right;">{result.overall_confidence:.0f}%</td>
                    <td style="text-align: center;">{result.total_discrepancies}</td>
                </tr>
                """
        
        # Build discrepancy breakdown
        disc_rows = ""
        for dtype, count in sorted(disc_types.items(), key=lambda x: x[1], reverse=True):
            disc_rows += f"<tr><td>{dtype.replace('_', ' ').title()}</td><td style='text-align: right;'>{count}</td></tr>"
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Provider Validation Report</title>
            <style>
                body {{ font-family: 'Segoe UI', Arial, sans-serif; margin: 40px; color: #333; }}
                .header {{ text-align: center; margin-bottom: 30px; border-bottom: 3px solid #007bff; padding-bottom: 20px; }}
                .header h1 {{ color: #007bff; margin: 0; }}
                .header p {{ color: #6c757d; margin: 5px 0; }}
                .stats-grid {{ display: flex; justify-content: space-around; margin: 30px 0; }}
                .stat-box {{ text-align: center; padding: 20px 40px; border-radius: 10px; }}
                .stat-value {{ font-size: 48px; font-weight: bold; }}
                .stat-label {{ font-size: 14px; color: #6c757d; }}
                .green {{ background-color: #d4edda; color: #155724; }}
                .yellow {{ background-color: #fff3cd; color: #856404; }}
                .red {{ background-color: #f8d7da; color: #721c24; }}
                .blue {{ background-color: #d1ecf1; color: #0c5460; }}
                table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
                th {{ background-color: #343a40; color: white; padding: 12px; text-align: left; }}
                td {{ padding: 10px; border-bottom: 1px solid #ddd; }}
                tr:hover {{ background-color: #f5f5f5; }}
                .section {{ margin: 30px 0; }}
                .section h2 {{ color: #343a40; border-bottom: 2px solid #007bff; padding-bottom: 10px; }}
                .footer {{ text-align: center; margin-top: 40px; padding-top: 20px; border-top: 1px solid #ddd; color: #6c757d; font-size: 12px; }}
                @media print {{ .page-break {{ page-break-before: always; }} }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>🏥 Provider Data Validation Report</h1>
                <p>Generated: {datetime.now().strftime('%B %d, %Y at %H:%M')}</p>
                <p>Report ID: {report.id[:8]}</p>
            </div>
            
            <div class="stats-grid">
                <div class="stat-box blue">
                    <div class="stat-value">{report.total_providers}</div>
                    <div class="stat-label">TOTAL PROVIDERS</div>
                </div>
                <div class="stat-box green">
                    <div class="stat-value">{status_counts['validated']}</div>
                    <div class="stat-label">AUTO-UPDATED ✓</div>
                </div>
                <div class="stat-box yellow">
                    <div class="stat-value">{status_counts['needs_review']}</div>
                    <div class="stat-label">NEEDS REVIEW ⚠</div>
                </div>
                <div class="stat-box red">
                    <div class="stat-value">{status_counts['urgent']}</div>
                    <div class="stat-label">URGENT REVIEW ✗</div>
                </div>
            </div>
            
            <div class="section">
                <h2>📊 Key Metrics</h2>
                <table>
                    <tr><td>Average Confidence Score</td><td style="text-align: right; font-weight: bold;">{report.average_confidence:.1f}%</td></tr>
                    <tr><td>Total Processing Time</td><td style="text-align: right; font-weight: bold;">{report.processing_time_seconds:.1f} seconds</td></tr>
                    <tr><td>Validation Success Rate</td><td style="text-align: right; font-weight: bold;">{((status_counts['validated'] + status_counts['needs_review']) / report.total_providers * 100):.1f}%</td></tr>
                    <tr><td>Auto-Update Rate</td><td style="text-align: right; font-weight: bold;">{(status_counts['validated'] / report.total_providers * 100):.1f}%</td></tr>
                </table>
            </div>
            
            <div class="section">
                <h2>🔍 Discrepancy Breakdown</h2>
                <table>
                    <tr><th>Discrepancy Type</th><th style="text-align: right;">Count</th></tr>
                    {disc_rows if disc_rows else "<tr><td colspan='2'>No discrepancies found</td></tr>"}
                </table>
            </div>
            
            <div class="section page-break">
                <h2>👥 Provider Details (Top 50)</h2>
                <table>
                    <tr>
                        <th>NPI</th>
                        <th>Provider Name</th>
                        <th>Specialty</th>
                        <th style="text-align: center;">Status</th>
                        <th style="text-align: right;">Confidence</th>
                        <th style="text-align: center;">Issues</th>
                    </tr>
                    {provider_rows}
                </table>
            </div>
            
            <div class="footer">
                <p>Provider Data Validation System - Automated Report</p>
                <p>© {datetime.now().year} Healthcare Data Solutions</p>
            </div>
        </body>
        </html>
        """
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        return str(filepath)
    
    def generate_summary_stats(
        self,
        providers: List[Provider],
        validation_results: Dict[str, ValidationResult]
    ) -> Dict[str, Any]:
        """
        Generate summary statistics for display in dashboard.
        """
        total = len(providers)
        
        auto_updated = sum(1 for p in providers if validation_results.get(p.id) and validation_results[p.id].auto_updated)
        needs_review = sum(1 for p in providers if validation_results.get(p.id) and validation_results[p.id].needs_review)
        urgent = sum(1 for p in providers if validation_results.get(p.id) and validation_results[p.id].urgent_review)
        pending = total - auto_updated - needs_review - urgent
        
        # Confidence distribution
        confidences = [
            validation_results[p.id].overall_confidence 
            for p in providers 
            if validation_results.get(p.id)
        ]
        
        # Discrepancy types
        disc_types = {}
        all_discrepancies = []
        for p in providers:
            result = validation_results.get(p.id)
            if result:
                all_discrepancies.extend(result.discrepancies)
                for disc in result.discrepancies:
                    dtype = disc.type.value
                    disc_types[dtype] = disc_types.get(dtype, 0) + 1
        
        # Priority breakdown
        priority_counts = {
            'high': sum(1 for d in all_discrepancies if d.priority == Priority.HIGH),
            'medium': sum(1 for d in all_discrepancies if d.priority == Priority.MEDIUM),
            'low': sum(1 for d in all_discrepancies if d.priority == Priority.LOW)
        }
        
        return {
            'total_providers': total,
            'auto_updated': auto_updated,
            'needs_review': needs_review,
            'urgent': urgent,
            'pending': pending,
            'average_confidence': sum(confidences) / len(confidences) if confidences else 0,
            'min_confidence': min(confidences) if confidences else 0,
            'max_confidence': max(confidences) if confidences else 0,
            'total_discrepancies': len(all_discrepancies),
            'discrepancy_types': disc_types,
            'priority_breakdown': priority_counts,
            'confidence_distribution': {
                'high (80-100%)': sum(1 for c in confidences if c >= 80),
                'medium (60-79%)': sum(1 for c in confidences if 60 <= c < 80),
                'low (<60%)': sum(1 for c in confidences if c < 60)
            }
        }
    
    def export_to_excel(
        self,
        providers: List[Provider],
        validation_results: Dict[str, ValidationResult],
        filename: Optional[str] = None
    ) -> str:
        """
        Export validation results to Excel format.
        
        Note: Requires openpyxl installed.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            # Fall back to CSV if openpyxl not available
            return self.generate_csv_report(providers, validation_results, filename)
        
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"validation_report_{timestamp}.xlsx"
        
        filepath = self.reports_dir / filename
        
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        stats = self.generate_summary_stats(providers, validation_results)
        
        summary_data = [
            ["Provider Data Validation Report"],
            [f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"],
            [""],
            ["Metric", "Value"],
            ["Total Providers", stats['total_providers']],
            ["Auto-Updated", stats['auto_updated']],
            ["Needs Review", stats['needs_review']],
            ["Urgent Review", stats['urgent']],
            ["Average Confidence", f"{stats['average_confidence']:.1f}%"],
            ["Total Discrepancies", stats['total_discrepancies']]
        ]
        
        for row in summary_data:
            ws_summary.append(row)
        
        # Provider details sheet
        ws_providers = wb.create_sheet("Providers")
        
        headers = ['NPI', 'Provider Name', 'Practice', 'Specialty', 'Phone', 
                   'City', 'State', 'Status', 'Confidence', 'Discrepancies']
        ws_providers.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color="343a40", end_color="343a40", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws_providers.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
        
        # Add provider data
        for provider in providers:
            result = validation_results.get(provider.id)
            
            status = "Pending"
            if result:
                if result.auto_updated:
                    status = "Auto-Updated"
                elif result.needs_review:
                    status = "Needs Review"
                elif result.urgent_review:
                    status = "Urgent"
            
            row = [
                provider.npi,
                provider.full_name(),
                provider.practice_name,
                provider.specialty,
                provider.contact.phone,
                provider.address.city,
                provider.address.state,
                status,
                f"{result.overall_confidence:.0f}%" if result else "N/A",
                result.total_discrepancies if result else 0
            ]
            ws_providers.append(row)
        
        # Discrepancies sheet
        ws_disc = wb.create_sheet("Discrepancies")
        disc_headers = ['Provider NPI', 'Type', 'Field', 'Current Value', 
                        'Validated Value', 'Priority', 'Confidence']
        ws_disc.append(disc_headers)
        
        for col, header in enumerate(disc_headers, 1):
            cell = ws_disc.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
        
        for provider in providers:
            result = validation_results.get(provider.id)
            if result:
                for disc in result.discrepancies:
                    row = [
                        provider.npi,
                        disc.type.value.replace('_', ' ').title(),
                        disc.field_name,
                        disc.current_value,
                        disc.validated_value,
                        disc.priority.value.upper(),
                        f"{disc.confidence:.0f}%"
                    ]
                    ws_disc.append(row)
        
        # Adjust column widths
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)
        return str(filepath)


# Singleton instance
report_generator = ReportGenerator()
